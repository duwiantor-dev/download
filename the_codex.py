import io
import re
import zipfile
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# APP CONFIG
# ============================================================
APP_TITLE = "The Codex"
MAX_MASS_FILES = 50
MAX_TOTAL_UPLOAD_MB = 200
BIGSELLER_MAX_ROWS_PER_FILE = 10000

st.set_page_config(page_title=APP_TITLE, layout="wide")


# ============================================================
# SESSION STATE
# ============================================================
SESSION_DEFAULTS = {
    "download_cache": {},
    "summary_cache": {},
    "stock_shopee_areas_loaded": [],
    "stock_tiktokshop_areas_loaded": [],
}
for _k, _v in SESSION_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ============================================================
# GENERIC HELPERS
# ============================================================
def s(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def s_clean(x) -> str:
    if x is None:
        return ""
    txt = str(x).replace("\xa0", " ")
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()


def su(x) -> str:
    return s_clean(x).upper()


def norm_sku(v) -> str:
    txt = su(v)
    if not txt:
        return ""
    if re.fullmatch(r"\d+\.0", txt):
        txt = txt[:-2]
    txt = re.sub(r"\s+", "", txt)
    return txt


def split_sku_addons(full_sku: str) -> Tuple[str, List[str]]:
    parts = [p.strip() for p in s_clean(full_sku).split("+") if p and s_clean(p)]
    if not parts:
        return "", []
    return parts[0], parts[1:]


def normalize_addon_code(x) -> str:
    return su(x)


def parse_number_like_id(x) -> str:
    if x is None:
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if pd.isna(x):
            return ""
        if x.is_integer():
            return str(int(x))
        return str(x)
    return s_clean(x)


def to_int_or_none(v) -> Optional[int]:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if pd.isna(v):
            return None
        return int(round(v))
    digits = re.findall(r"\d+", s_clean(v))
    if not digits:
        return None
    return int("".join(digits))


def parse_price_cell(val) -> Optional[int]:
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        try:
            if isinstance(val, float) and pd.isna(val):
                return None
            return int(round(float(val)))
        except Exception:
            return None

    txt = s_clean(val)
    if not txt:
        return None
    txt = txt.replace("Rp", "").replace("rp", "").replace(" ", "")

    if "." in txt and "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "." in txt and "," not in txt:
        txt = txt.replace(".", "")
    elif "," in txt and "." not in txt:
        txt = txt.replace(",", "")

    try:
        return int(round(float(txt)))
    except Exception:
        return None


def apply_multiplier_if_needed(x: Optional[int], threshold: int = 1_000_000, multiplier: int = 1000) -> int:
    if x is None:
        return 0
    if x < threshold:
        return int(x) * multiplier
    return int(x)


def total_upload_size_mb(files: List[Any]) -> float:
    total = 0
    for f in files:
        try:
            total += len(f.getvalue())
        except Exception:
            pass
    return total / (1024 * 1024)


def lower_map_headers(ws: Worksheet, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = s_clean(v).lower()
        if key and key not in m:
            m[key] = c
    return m


def get_header_col_fuzzy(ws: Worksheet, header_row: int, candidates: List[str]) -> Optional[int]:
    m = lower_map_headers(ws, header_row)
    normalized = {re.sub(r"[^a-z0-9]", "", k): v for k, v in m.items()}
    for cand in candidates:
        target = re.sub(r"[^a-z0-9]", "", s_clean(cand).lower())
        if target in normalized:
            return normalized[target]
    return None


def find_header_row_by_exact(ws: Worksheet, header_text: str, scan_rows: int = 150) -> Optional[int]:
    target = su(header_text)
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        for c in range(1, ws.max_column + 1):
            if su(ws.cell(r, c).value) == target:
                return r
    return None


def find_row_contains(ws: Worksheet, needle: str, scan_rows: int = 300) -> Optional[int]:
    target = su(needle)
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        for c in range(1, ws.max_column + 1):
            v = su(ws.cell(r, c).value)
            if v and (target == v or target in v):
                return r
    return None


def get_first_sheet(wb) -> Worksheet:
    return wb[wb.sheetnames[0]]


def build_merged_lookup_map(ws: Worksheet) -> Dict[Tuple[int, int], object]:
    merged_map: Dict[Tuple[int, int], object] = {}
    for mr in ws.merged_cells.ranges:
        top_left_val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_map[(r, c)] = top_left_val
    return merged_map


def get_cell_or_merged_value(ws: Worksheet, merged_map: Dict[Tuple[int, int], object], row: int, col: int):
    v = ws.cell(row, col).value
    if v not in (None, ""):
        return v
    return merged_map.get((row, col))


def safe_set_cell_value(ws: Worksheet, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        coord = cell.coordinate
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return
        return
    cell.value = value


def workbook_to_bytes(wb) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def zip_named_files(named_files: List[Tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for fname, fbytes in named_files:
            zf.writestr(fname, fbytes)
    return buf.getvalue()


def make_issues_workbook(issues: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "issues_report"
    headers = ["file", "row", "sku_full", "old_value", "new_value", "reason"]
    ws.append(headers)
    for item in issues:
        ws.append([
            item.get("file", ""),
            item.get("row", ""),
            item.get("sku_full", ""),
            item.get("old_value", ""),
            item.get("new_value", ""),
            item.get("reason", ""),
        ])
    return workbook_to_bytes(wb)


def render_summary(title: str, summary: Dict[str, Any]):
    st.subheader(title)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Jumlah File", int(summary.get("files_total", 0)))
    c2.metric("Baris Diperiksa", int(summary.get("rows_scanned", 0)))
    c3.metric("Baris Diupdate", int(summary.get("rows_written", 0)))
    c4.metric("Skip / Issue", int(summary.get("rows_unmatched", 0) + summary.get("issues_count", 0)))


def cache_downloads(
    cache_key: str,
    result_name: str,
    result_bytes: Optional[bytes],
    issues_bytes: Optional[bytes],
    summary: Optional[Dict[str, Any]] = None,
    issues_name: str = "issues_report.xlsx",
):
    st.session_state.download_cache[cache_key] = {
        "result_name": result_name,
        "result_bytes": result_bytes,
        "issues_name": issues_name,
        "issues_bytes": issues_bytes,
    }
    if summary is not None:
        st.session_state.summary_cache[cache_key] = summary


def render_downloads(cache_key: str):
    payload = st.session_state.download_cache.get(cache_key)
    if not payload:
        return
    if payload.get("result_bytes"):
        st.download_button(
            "Download Hasil",
            payload["result_bytes"],
            file_name=payload["result_name"],
            key=f"dl_{cache_key}_result",
        )
    if payload.get("issues_bytes"):
        st.download_button(
            "Download Issues",
            payload["issues_bytes"],
            file_name=payload["issues_name"],
            key=f"dl_{cache_key}_issues",
        )


def render_cached_summary(cache_key: str, title: str = "Ringkasan Hasil"):
    summary = st.session_state.summary_cache.get(cache_key)
    if summary:
        render_summary(title, summary)


def get_change_sheet(wb):
    for sname in wb.sheetnames:
        if su(sname) == "CHANGE":
            return wb[sname]
    raise ValueError("Sheet 'CHANGE' tidak ditemukan di Pricelist.")


def find_header_row_by_candidates(
    ws: Worksheet,
    required_candidates: Dict[str, List[str]],
    scan_rows: int = 10
) -> Tuple[int, Dict[str, int]]:
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        found: Dict[str, int] = {}
        ok = True
        for key, candidates in required_candidates.items():
            col = get_header_col_fuzzy(ws, r, candidates)
            if col is None:
                ok = False
                break
            found[key] = col
        if ok:
            return r, found
    raise ValueError("Header tidak ditemukan. Pastikan file memiliki kolom yang dibutuhkan.")


# ============================================================
# STOCK PRICELIST HELPERS
# ============================================================
def sheet_range_between(sheetnames: List[str], start: str, end: str) -> List[str]:
    up = [su(x) for x in sheetnames]
    if start.upper() not in up or end.upper() not in up:
        raise ValueError(f"Sheet range tidak valid. Pastikan ada '{start}' dan '{end}'.")
    i0 = up.index(start.upper())
    i1 = up.index(end.upper())
    if i0 > i1:
        i0, i1 = i1, i0
    return sheetnames[i0:i1 + 1]


def delete_coming_block_in_laptop(ws: Worksheet):
    r_start = find_row_contains(ws, "COMING", scan_rows=600)
    r_end = find_row_contains(ws, "END COMING", scan_rows=1200)
    if r_start and r_end and r_end >= r_start:
        ws.delete_rows(r_start, r_end - r_start + 1)


def find_tot_col(ws: Worksheet, header_row_hint: int) -> Tuple[int, int]:
    for c in range(1, ws.max_column + 1):
        if su(ws.cell(header_row_hint, c).value) == "TOT":
            return header_row_hint, c
    for r in range(1, min(12, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if su(ws.cell(r, c).value) == "TOT":
                return r, c
    raise ValueError("Kolom 'TOT' tidak ketemu.")


def build_area_meta(ws: Worksheet, merged_map: Dict[Tuple[int, int], object], area_row: int, start_col: int) -> Dict[int, str]:
    col_area: Dict[int, str] = {}
    for c in range(start_col, ws.max_column + 1):
        area_raw = get_cell_or_merged_value(ws, merged_map, area_row, c)
        area_name = su(area_raw)
        if area_name:
            col_area[c] = area_name
    return col_area


def build_stock_lookup_from_sheet_fast(ws: Worksheet, sheet_name: str):
    area_row = 3
    header_row = find_header_row_by_exact(ws, "KODEBARANG", scan_rows=200)
    if header_row is None:
        header_row = find_header_row_by_exact(ws, "KODE BARANG", scan_rows=200)
    if header_row is None:
        raise ValueError(f"[{sheet_name}] Header 'KODEBARANG' tidak ketemu.")

    sku_col = None
    for c in range(1, ws.max_column + 1):
        v = su(ws.cell(header_row, c).value)
        if v in ("KODEBARANG", "KODE BARANG"):
            sku_col = c
            break
    if sku_col is None:
        raise ValueError(f"[{sheet_name}] Kolom 'KODEBARANG' / 'KODE BARANG' tidak ditemukan.")

    header_row_used, tot_col = find_tot_col(ws, header_row)
    merged_map = build_merged_lookup_map(ws)
    col_area = build_area_meta(ws, merged_map, area_row=area_row, start_col=tot_col + 1)

    sku_map: Dict[str, Dict[str, Any]] = {}
    areas: Set[str] = set(col_area.values())
    for r in range(header_row_used + 1, ws.max_row + 1):
        sku = s_clean(ws.cell(r, sku_col).value)
        if not sku:
            continue
        sku_key = norm_sku(sku)
        if sku_key in ("TOTAL", "KODEBARANG", "KODE BARANG", "KODEBARANG."):
            continue

        tot_val = to_int_or_none(ws.cell(r, tot_col).value)
        by_area: Dict[str, int] = {}
        for c, area_name in col_area.items():
            v = to_int_or_none(ws.cell(r, c).value)
            if v is None:
                continue
            by_area[area_name] = by_area.get(area_name, 0) + int(v)
        sku_map[sku_key] = {"TOT": tot_val, "by_area": by_area}
    return sku_map, sorted(areas)


def build_stock_lookup_from_pricelist_bytes(pl_bytes: bytes):
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True, read_only=False)
    for sname in wb.sheetnames:
        if su(sname) == "LAPTOP":
            delete_coming_block_in_laptop(wb[sname])
            break
    target_sheets = sheet_range_between(wb.sheetnames, "LAPTOP", "SER OTH CON")
    merged_lookup: Dict[str, Dict[str, Any]] = {}
    areas_all: Set[str] = set()
    for sname in target_sheets:
        sku_map, areas = build_stock_lookup_from_sheet_fast(wb[sname], sname)
        merged_lookup.update(sku_map)
        areas_all |= set(areas)
    if not merged_lookup:
        raise ValueError("Pricelist terbaca, tapi lookup stok kosong.")
    return merged_lookup, sorted(areas_all)


def pick_stock_value(sku_full: str, stock_lookup: Dict[str, Dict], mode: str, chosen_areas: Set[str]) -> Optional[int]:
    base, _ = split_sku_addons(sku_full)
    base_key = norm_sku(base)
    if not base_key or base_key not in stock_lookup:
        return None
    rec = stock_lookup[base_key]
    tot = rec.get("TOT")
    by_area = rec.get("by_area", {}) or {}
    if mode == "Stok Nasional (TOT)":
        return tot if tot is not None else None
    if mode == "Stok Area":
        if not chosen_areas:
            return None
        total = 0
        hit = False
        for area_name, v in by_area.items():
            if area_name in chosen_areas:
                total += int(v)
                hit = True
        return total if hit else None
    return None


# ============================================================
# STOCK PROCESSORS
# ============================================================
def find_shopee_columns_readonly(ws) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 7
    row_vals = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    sku_col = None
    qty_col = None
    for idx, val in enumerate(row_vals, start=1):
        v = su(val)
        if v == "SKU":
            sku_col = idx
        if v == "STOK":
            qty_col = idx
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/Stok tidak ketemu pada template Shopee (Mall & Star).")
    return data_start, sku_col, qty_col


def find_shopee_columns_normal(ws: Worksheet) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 7
    sku_col = None
    qty_col = None
    for c in range(1, ws.max_column + 1):
        v = su(ws.cell(header_row, c).value)
        if v == "SKU":
            sku_col = c
        if v == "STOK":
            qty_col = c
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/Stok tidak ketemu pada template Shopee (Mall & Star).")
    return data_start, sku_col, qty_col


def collect_changed_rows_stock_shopee(file_bytes: bytes, stock_lookup: Dict[str, Dict], mode: str, chosen_areas: Set[str]):
    stats = {"rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0}
    changed_rows: List[List[Any]] = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    data_start, sku_col, qty_col = find_shopee_columns_readonly(ws)

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        row_list = list(row)
        sku_full = s_clean(row_list[sku_col - 1] if len(row_list) >= sku_col else None)
        if not sku_full:
            continue
        stats["rows_scanned"] += 1
        old_qty = to_int_or_none(row_list[qty_col - 1] if len(row_list) >= qty_col else None)
        new_qty = pick_stock_value(sku_full, stock_lookup, mode, chosen_areas)
        if new_qty is None:
            stats["rows_unmatched"] += 1
            continue
        if old_qty is not None and int(old_qty) == int(new_qty):
            stats["rows_unchanged"] += 1
            continue
        if len(row_list) < qty_col:
            row_list.extend([None] * (qty_col - len(row_list)))
        row_list[qty_col - 1] = int(new_qty)
        changed_rows.append(row_list)
        stats["rows_written"] += 1
    wb.close()
    return changed_rows, stats


def write_stock_shopee_output(template_bytes: bytes, changed_rows_all: List[List[Any]]) -> bytes:
    out_wb = load_workbook(io.BytesIO(template_bytes))
    out_ws = get_first_sheet(out_wb)
    data_start, _, _ = find_shopee_columns_normal(out_ws)
    if out_ws.max_row >= data_start:
        out_ws.delete_rows(data_start, out_ws.max_row - data_start + 1)
    for idx, row_vals in enumerate(changed_rows_all, start=data_start):
        for c, val in enumerate(row_vals, start=1):
            out_ws.cell(idx, c).value = val
    return workbook_to_bytes(out_wb)


def process_stock_shopee(mass_files: List[Any], pricelist_file: Any, mode: str, chosen_areas: Set[str]):
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
    changed_rows_all: List[List[Any]] = []
    issues: List[Dict[str, Any]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        try:
            rows, stats = collect_changed_rows_stock_shopee(mf.getvalue(), stock_lookup, mode, chosen_areas)
            changed_rows_all.extend(rows)
            for k in ("rows_scanned", "rows_written", "rows_unchanged", "rows_unmatched"):
                summary[k] += stats[k]
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})

    if summary["rows_written"] == 0 and not issues:
        issues.append({"file": "", "reason": "Tidak ada baris berubah / tidak ada SKU yang match."})

    result_bytes = write_stock_shopee_output(mass_files[0].getvalue(), changed_rows_all)
    summary["issues_count"] = len(issues)
    return result_bytes, make_issues_workbook(issues) if issues else None, summary


def find_tiktokshop_columns_readonly(ws) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 6
    qty_headers = {"KUANTITAS", "JUMLAH DI SHOP LOCATION", "QUANTITY"}
    row_vals = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    sku_col = None
    qty_col = None
    for idx, val in enumerate(row_vals, start=1):
        v = su(val)
        if v in ("SKU PENJUAL", "SELLER SKU"):
            sku_col = idx
        if v in qty_headers:
            qty_col = idx
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/stok tidak ketemu pada template TikTokShop.")
    return data_start, sku_col, qty_col


def find_tiktokshop_columns_normal(ws: Worksheet) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 6
    qty_headers = {"KUANTITAS", "JUMLAH DI SHOP LOCATION", "QUANTITY"}
    sku_col = None
    qty_col = None
    for c in range(1, ws.max_column + 1):
        v = su(ws.cell(header_row, c).value)
        if v in ("SKU PENJUAL", "SELLER SKU"):
            sku_col = c
        if v in qty_headers:
            qty_col = c
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/stok tidak ketemu pada template TikTokShop.")
    return data_start, sku_col, qty_col


def collect_changed_rows_stock_tiktokshop(file_bytes: bytes, stock_lookup: Dict[str, Dict], mode: str, chosen_areas: Set[str]):
    stats = {"rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0}
    changed_rows: List[List[Any]] = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    data_start, sku_col, qty_col = find_tiktokshop_columns_readonly(ws)

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        row_list = list(row)
        sku_full = s_clean(row_list[sku_col - 1] if len(row_list) >= sku_col else None)
        if not sku_full:
            continue
        stats["rows_scanned"] += 1
        old_qty = to_int_or_none(row_list[qty_col - 1] if len(row_list) >= qty_col else None)
        new_qty = pick_stock_value(sku_full, stock_lookup, mode, chosen_areas)
        if new_qty is None:
            stats["rows_unmatched"] += 1
            continue
        if old_qty is not None and int(old_qty) == int(new_qty):
            stats["rows_unchanged"] += 1
            continue
        if len(row_list) < qty_col:
            row_list.extend([None] * (qty_col - len(row_list)))
        row_list[qty_col - 1] = int(new_qty)
        changed_rows.append(row_list)
        stats["rows_written"] += 1
    wb.close()
    return changed_rows, stats


def write_stock_tiktokshop_output(template_bytes: bytes, changed_rows_all: List[List[Any]]) -> bytes:
    out_wb = load_workbook(io.BytesIO(template_bytes))
    out_ws = get_first_sheet(out_wb)
    data_start, _, _ = find_tiktokshop_columns_normal(out_ws)
    if out_ws.max_row >= data_start:
        out_ws.delete_rows(data_start, out_ws.max_row - data_start + 1)
    for idx, row_vals in enumerate(changed_rows_all, start=data_start):
        for c, val in enumerate(row_vals, start=1):
            out_ws.cell(idx, c).value = val
    return workbook_to_bytes(out_wb)


def process_stock_tiktokshop(mass_files: List[Any], pricelist_file: Any, mode: str, chosen_areas: Set[str]):
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
    changed_rows_all: List[List[Any]] = []
    issues: List[Dict[str, Any]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        try:
            rows, stats = collect_changed_rows_stock_tiktokshop(mf.getvalue(), stock_lookup, mode, chosen_areas)
            changed_rows_all.extend(rows)
            for k in ("rows_scanned", "rows_written", "rows_unchanged", "rows_unmatched"):
                summary[k] += stats[k]
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})

    if summary["rows_written"] == 0 and not issues:
        issues.append({"file": "", "reason": "Tidak ada baris berubah / tidak ada SKU yang match."})

    result_bytes = write_stock_tiktokshop_output(mass_files[0].getvalue(), changed_rows_all)
    summary["issues_count"] = len(issues)
    return result_bytes, make_issues_workbook(issues) if issues else None, summary


# ============================================================
# PRICE LOADERS
# ============================================================
def load_addon_map_generic(addon_bytes: bytes) -> Dict[str, int]:
    wb = load_workbook(io.BytesIO(addon_bytes), data_only=True)
    ws = wb.active
    code_candidates = ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON", "Standarisasi Kode SKU di Varian"]
    price_candidates = ["harga", "HARGA", "Price", "PRICE", "Harga"]

    header_row = None
    code_col = None
    price_col = None
    for r in range(1, 30):
        code_col = get_header_col_fuzzy(ws, r, code_candidates)
        price_col = get_header_col_fuzzy(ws, r, price_candidates)
        if code_col and price_col:
            header_row = r
            break
    if header_row is None or code_col is None or price_col is None:
        raise ValueError("Header Addon Mapping tidak ketemu. Pastikan ada kolom addon_code & harga (atau setara).")

    addon_map: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = normalize_addon_code(ws.cell(row=r, column=code_col).value)
        if not code:
            continue
        price_raw = parse_price_cell(ws.cell(row=r, column=price_col).value)
        if price_raw is None:
            continue
        addon_map[code] = int(apply_multiplier_if_needed(price_raw))
    return addon_map


def find_header_row_and_cols_pricelist_fixed(ws: Worksheet, required_price_cols: List[str]) -> Tuple[int, int, Dict[str, int]]:
    sku_candidates = ["KODEBARANG", "KODE BARANG", "SKU", "SKU NO", "SKU_NO", "KODEBARANG "]

    for header_row in range(1, min(25, ws.max_row) + 1):
        sku_col = get_header_col_fuzzy(ws, header_row, sku_candidates)
        if sku_col is None:
            continue

        price_cols: Dict[str, int] = {}
        all_found = True
        for p in required_price_cols:
            col = get_header_col_fuzzy(ws, header_row, [p])
            if col is None:
                all_found = False
                break
            price_cols[p] = col

        if all_found:
            return header_row, sku_col, price_cols

    raise ValueError(
        f"Header Pricelist tidak ketemu. Pastikan ada kolom SKU/KODEBARANG dan kolom harga {required_price_cols}."
    )


def load_pricelist_price_map(pl_bytes: bytes, needed_cols: List[str]) -> Dict[str, Dict[str, int]]:
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True)
    ws = get_change_sheet(wb)
    header_row, sku_col, price_cols = find_header_row_and_cols_pricelist_fixed(ws, needed_cols)
    result: Dict[str, Dict[str, int]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        sku = norm_sku(ws.cell(row=r, column=sku_col).value)
        if not sku:
            continue
        result[sku] = {}
        for label, col in price_cols.items():
            raw = parse_price_cell(ws.cell(row=r, column=col).value)
            if raw is not None:
                result[sku][label] = int(apply_multiplier_if_needed(raw))
    return result


def compute_price_from_maps(sku_full: str, price_map: Dict[str, Dict[str, int]], addon_map: Dict[str, int], price_key: str, discount_rp: int) -> Tuple[Optional[int], str]:
    base_sku, addons = split_sku_addons(sku_full)
    base_sku = norm_sku(base_sku)
    if not base_sku:
        return None, "SKU kosong"
    pl = price_map.get(base_sku)
    if not pl:
        return None, f"Base SKU '{base_sku}' tidak ada di Pricelist"
    base_price = pl.get(price_key)
    if base_price is None:
        return None, f"Harga {price_key} kosong di Pricelist untuk SKU '{base_sku}'"
    addon_total = 0
    for addon in addons:
        code = normalize_addon_code(addon)
        if code and code not in addon_map:
            return None, f"Addon '{code}' tidak ada di file Addon Mapping"
        addon_total += int(addon_map.get(code, 0))
    final_price = int(base_price) + addon_total - int(discount_rp)
    if final_price < 0:
        final_price = 0
    return final_price, f"{price_key} + addon - diskon"


# ============================================================
# PRICE PROCESSORS
# ============================================================
def process_shopee_price_files(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, price_key: str, page_title: str, mode: str):
    price_map = load_pricelist_price_map(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active

        if mode == "normal":
            header_row, found_cols = find_header_row_by_candidates(
                ws,
                {
                    "sku": ["SKU", "SKU Ref. No.(Optional)", "SKU Ref No Optional", "SKU Penjual"],
                    "price": ["Harga", "Harga Diskon", "Harga diskon"],
                },
                scan_rows=10,
            )
        else:
            header_row, found_cols = find_header_row_by_candidates(
                ws,
                {
                    "sku": ["SKU", "SKU Ref. No.(Optional)", "SKU Ref No Optional", "SKU Penjual"],
                    "price": ["Harga Diskon", "Harga diskon", "Harga"],
                },
                scan_rows=10,
            )

        sku_col = found_cols["sku"]
        price_col = found_cols["price"]
        data_start_fixed = header_row + 1

        changed_rows: List[int] = []
        for r in range(data_start_fixed, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue

            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, price_key, discount_rp)

            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({
                    "file": mf.name,
                    "row": r,
                    "sku_full": sku_full,
                    "old_value": old_price,
                    "new_value": "",
                    "reason": reason,
                })
                continue

            if old_price is not None and int(old_price) == int(new_price):
                continue

            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, data_start_fixed - 1, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})

        output_files.append((f"hasil_{page_title.lower().replace(' ', '_')}_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), f"hasil_{page_title.lower().replace(' ', '_')}.zip", make_issues_workbook(issues) if issues else None, summary


def process_tiktokshop_price_normal(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int):
    price_map = load_pricelist_price_map(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        sku_col = get_header_col_fuzzy(ws, 3, ["SKU Penjual", "Seller SKU"])
        price_col = get_header_col_fuzzy(ws, 3, ["Harga Ritel (Mata Uang Lokal)", "Harga", "Price"])
        if sku_col is None or price_col is None:
            issues.append({"file": mf.name, "reason": "Header mass update TikTokShop tidak sesuai."})
            continue

        changed_rows: List[int] = []
        for r in range(6, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, "M3", discount_rp)
            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_price, "new_value": "", "reason": reason})
                continue
            if old_price is not None and int(old_price) == int(new_price):
                continue
            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, 5, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})
        output_files.append((f"hasil_harga_normal_tiktokshop_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), "hasil_harga_normal_tiktokshop.zip", make_issues_workbook(issues) if issues else None, summary


def process_powemerchant_price_files(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, page_title: str):
    price_map = load_pricelist_price_map(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        sku_col = get_header_col_fuzzy(ws, 3, ["SKU Penjual", "Seller SKU"])
        price_col = get_header_col_fuzzy(ws, 3, ["Harga Ritel (Mata Uang Lokal)", "Harga", "Price"])
        if sku_col is None or price_col is None:
            issues.append({"file": mf.name, "reason": "Header mass update PowerMerchant tidak sesuai."})
            continue

        changed_rows: List[int] = []
        for r in range(6, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, "M4", discount_rp)
            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_price, "new_value": "", "reason": reason})
                continue
            if old_price is not None and int(old_price) == int(new_price):
                continue
            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, 5, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})
        output_files.append((f"hasil_{page_title.lower().replace(' ', '_')}_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), f"hasil_{page_title.lower().replace(' ', '_')}.zip", make_issues_workbook(issues) if issues else None, summary


def process_tiktokshop_price_coret(input_file: Any, pricelist_file: Any, addon_file: Any, discount_rp: int, only_changed: bool = True):
    price_map = load_pricelist_price_map(pricelist_file.getvalue(), ["M3"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    wb_in = load_workbook(io.BytesIO(input_file.getvalue()), data_only=True)
    ws_in = wb_in.active

    out_wb = Workbook()
    ws_out = out_wb.active
    ws_out.title = "Sheet1"
    headers = [
        "Product_id (wajib)",
        "SKU_id (wajib)",
        "Harga Penawaran (wajib)",
        "Total Stok Promosi (opsional)\n1. Total Stok Promosi≤ Stok \n2. Jika tidak diisi artinya tidak terbatas",
        "Batas Pembelian (opsional)\n1. 1 ≤ Batas pembelian≤ 99\n2. Jika tidak diisi artinya tidak terbatas",
    ]
    for i, h in enumerate(headers, start=1):
        ws_out.cell(row=1, column=i).value = h

    issues: List[Dict[str, Any]] = []
    summary = {"files_total": 1, "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}
    row_out = 2

    for r in range(6, ws_in.max_row + 1):
        product_id = parse_number_like_id(ws_in.cell(row=r, column=1).value)
        sku_id = parse_number_like_id(ws_in.cell(row=r, column=4).value)
        old_price = parse_price_cell(ws_in.cell(row=r, column=6).value)
        stock = to_int_or_none(ws_in.cell(row=r, column=7).value)
        seller_sku = s_clean(ws_in.cell(row=r, column=8).value or ws_in.cell(row=r, column=5).value)
        if not seller_sku:
            continue
        summary["rows_scanned"] += 1
        new_price, reason = compute_price_from_maps(seller_sku, price_map, addon_map, "M3", discount_rp)
        if new_price is None:
            summary["rows_unmatched"] += 1
            issues.append({"file": input_file.name, "row": r, "sku_full": seller_sku, "old_value": old_price, "new_value": "", "reason": reason})
            continue
        if only_changed and old_price is not None and int(old_price) == int(new_price):
            continue
        ws_out.cell(row=row_out, column=1).value = product_id
        ws_out.cell(row=row_out, column=2).value = sku_id
        ws_out.cell(row=row_out, column=3).value = new_price
        ws_out.cell(row=row_out, column=4).value = stock if stock is not None else ""
        row_out += 1
        summary["rows_written"] += 1

    summary["issues_count"] = len(issues)
    return workbook_to_bytes(out_wb), "hasil_harga_coret_tiktokshop.xlsx", make_issues_workbook(issues) if issues else None, summary


def process_bigseller(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int):
    price_map = load_pricelist_price_map(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}
    output_parts: List[Tuple[str, bytes]] = []
    current_rows: List[List[Any]] = []
    current_part = 1
    output_header: List[Any] = []
    header_len = 0

    def flush_part():
        nonlocal current_rows, current_part, output_parts, output_header, header_len
        if not current_rows:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for c, val in enumerate(output_header, start=1):
            ws.cell(row=1, column=c).value = val
        for r_idx, row_vals in enumerate(current_rows, start=2):
            for c, val in enumerate(row_vals, start=1):
                ws.cell(row=r_idx, column=c).value = val
        output_parts.append((f"hasil_harga_normal_bigseller_part_{current_part}.xlsx", workbook_to_bytes(wb)))
        current_rows = []
        current_part += 1

    for mf in mass_files:
        wb = None
        try:
            wb = load_workbook(io.BytesIO(mf.getvalue()), read_only=False, data_only=False)
            ws = wb.worksheets[0]

            header_row, found_cols = find_header_row_by_candidates(
                ws,
                {
                    "sku": ["SKU", "Seller SKU", "SKU Penjual"],
                    "price": ["Price", "Harga", "Harga Jual"],
                },
                scan_rows=10,
            )

            sku_col = found_cols["sku"]
            harga_col = found_cols["price"]

            if not output_header:
                output_header = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
                header_len = ws.max_column

            mp = "M3" if "tiktok" in mf.name.lower() else "M4"

            for r in range(header_row + 1, ws.max_row + 1):
                sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
                if not sku_full:
                    continue

                summary["rows_scanned"] += 1
                old_price = parse_price_cell(ws.cell(row=r, column=harga_col).value)
                new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, mp, discount_rp)

                if new_price is None:
                    summary["rows_unmatched"] += 1
                    issues.append({
                        "file": mf.name,
                        "row": r,
                        "sku_full": sku_full,
                        "old_value": old_price,
                        "new_value": "",
                        "reason": reason,
                    })
                    continue

                if old_price is not None and int(old_price) == int(new_price):
                    continue

                row_vals = [ws.cell(row=r, column=c).value for c in range(1, header_len + 1)]
                row_vals[harga_col - 1] = new_price
                current_rows.append(row_vals)
                summary["rows_written"] += 1

                if len(current_rows) >= BIGSELLER_MAX_ROWS_PER_FILE:
                    flush_part()

        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

    flush_part()
    summary["issues_count"] = len(issues)
    if not output_parts:
        empty_wb = Workbook()
        empty_ws = empty_wb.active
        empty_ws.title = "Sheet1"
        if output_header:
            for c, val in enumerate(output_header, start=1):
                empty_ws.cell(row=1, column=c).value = val
        output_parts.append(("hasil_harga_normal_bigseller_part_1.xlsx", workbook_to_bytes(empty_wb)))
    if len(output_parts) == 1:
        return output_parts[0][1], output_parts[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_parts), "hasil_harga_normal_bigseller.zip", make_issues_workbook(issues) if issues else None, summary


# ============================================================
# SUBMIT CAMPAIGN PROCESSORS
# ============================================================
def process_submit_campaign_tiktokshop(
    mass_files: List[Any],
    pricelist_file: Any,
    addon_file: Any,
    discount_rp: int,
    price_key: str,
):
    price_map = load_pricelist_price_map(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())

    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {
        "files_total": len(mass_files),
        "rows_scanned": 0,
        "rows_written": 0,
        "rows_unmatched": 0,
        "issues_count": 0,
    }

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active

        header_row = 2
        data_start = 3

        sku_col = get_header_col_fuzzy(ws, header_row, [
            "SKU Name",
            "Nama SKU",
            "Seller SKU",
            "SKU Penjual",
        ])
        price_col = get_header_col_fuzzy(ws, header_row, [
            "Campaign price",
            "Campaign Price",
            "Harga Campaign",
            "Harga Promo",
        ])

        if sku_col is None or price_col is None:
            issues.append({
                "file": mf.name,
                "reason": "Header Submit Campaign tidak sesuai. Pastikan row 2 berisi header SKU dan Campaign price.",
            })
            continue

        changed_rows: List[int] = []

        for r in range(data_start, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue

            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(
                sku_full, price_map, addon_map, price_key, discount_rp
            )

            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({
                    "file": mf.name,
                    "row": r,
                    "sku_full": sku_full,
                    "old_value": old_price,
                    "new_value": "",
                    "reason": reason,
                })
                continue

            if old_price is not None and int(old_price) == int(new_price):
                continue

            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, data_start - 1, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({
                "file": mf.name,
                "reason": "Tidak ada baris berubah pada file ini.",
            })

        output_files.append((f"hasil_submit_campaign_tiktokshop_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)

    if len(output_files) == 1:
        return (
            output_files[0][1],
            output_files[0][0],
            make_issues_workbook(issues) if issues else None,
            summary,
        )

    return (
        zip_named_files(output_files),
        "hasil_submit_campaign_tiktokshop.zip",
        make_issues_workbook(issues) if issues else None,
        summary,
    )


# ============================================================
# UI HELPERS
# ============================================================
def page_header(title: str, desc: str, requirements: List[str]):
    st.title(title)
    st.caption(desc)
    with st.expander("Kebutuhan File", expanded=True):
        for item in requirements:
            st.write(f"- {item}")


def validate_mass_uploads(mass_files: List[Any]) -> Optional[str]:
    if not mass_files:
        return "Upload file mass update minimal 1 file."
    if len(mass_files) > MAX_MASS_FILES:
        return f"Maksimal {MAX_MASS_FILES} file per proses."
    if total_upload_size_mb(mass_files) > MAX_TOTAL_UPLOAD_MB:
        return f"Total upload melebihi {MAX_TOTAL_UPLOAD_MB} MB."
    return None


# ============================================================
# PAGES
# ============================================================
def render_dashboard():
    st.title(APP_TITLE)
    st.markdown("Aplikasi all-in-one untuk **Update Stok**, **Harga Normal**, **Harga Coret**, dan **Submit Campaign** marketplace.")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.subheader("Update Stok")
        st.write("- Shopee (Mall & Star)\n- TikTokShop")
    with c2:
        st.subheader("Update Harga Normal")
        st.write("- Shopee (Mall & Star)\n- TikTokShop\n- PowerMerchant\n- Bigseller")
    with c3:
        st.subheader("Update Harga Coret")
        st.write("- Shopee (Mall & Star)\n- TikTokShop\n- PowerMerchant")
    with c4:
        st.subheader("Submit Campaign")
        st.write("- TikTokShop\n- Shopee (Coming Soon)")
    st.info("Gunakan menu di sidebar untuk memilih fitur.")


def render_update_stok_shopee():
    page_header(
        "Update Stok Shopee (Mall & Star)",
        "Memproses file mass update Shopee (Mall & Star) berdasarkan stok dari pricelist multi-sheet.",
        ["Mass Update Shopee (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Jika pakai stok area, klik 'Load Data Area' dulu lalu pilih area"],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Shopee (Mall & Star)", type=["xlsx"], accept_multiple_files=True, key="stock_shopee_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_shopee_pl")
    mode = st.radio("Mode Stok", ["Stok Nasional (TOT)", "Stok Area"], horizontal=True)

    if st.button("Load Data Area", key="load_area_shopee"):
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
        else:
            try:
                _, areas = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
                st.session_state.stock_shopee_areas_loaded = areas
                st.success(f"Data area berhasil dimuat: {len(areas)} area")
            except Exception as e:
                st.error(f"Gagal load data area: {e}")

    chosen_areas = set()
    areas = st.session_state.stock_shopee_areas_loaded
    if mode == "Stok Area":
        chosen_areas = set(st.multiselect("Pilih Area", areas, key="stock_shopee_areas"))

    process_disabled = mode == "Stok Area" and (not areas or not chosen_areas)

    if st.button("Proses", key="btn_stock_shopee", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, issues_bytes, summary = process_stock_shopee(mass_files, pricelist_file, mode, chosen_areas)
            cache_downloads("stock_shopee", "hasil_update_stok_shopee.xlsx", result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_shopee")
    render_downloads("stock_shopee")


def render_update_stok_tiktokshop():
    page_header(
        "Update Stok TikTokShop",
        "Memproses file mass update TikTokShop berdasarkan stok dari pricelist multi-sheet.",
        ["Mass Update TikTokShop (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Jika pakai stok area, klik 'Load Data Area' dulu lalu pilih area"],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update TikTokShop", type=["xlsx"], accept_multiple_files=True, key="stock_tiktokshop_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_tiktokshop_pl")
    mode = st.radio("Mode Stok", ["Stok Nasional (TOT)", "Stok Area"], horizontal=True, key="stock_tiktokshop_mode")

    if st.button("Load Data Area", key="load_area_tiktokshop"):
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
        else:
            try:
                _, areas = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
                st.session_state.stock_tiktokshop_areas_loaded = areas
                st.success(f"Data area berhasil dimuat: {len(areas)} area")
            except Exception as e:
                st.error(f"Gagal load data area: {e}")

    chosen_areas = set()
    areas = st.session_state.stock_tiktokshop_areas_loaded
    if mode == "Stok Area":
        chosen_areas = set(st.multiselect("Pilih Area", areas, key="stock_tiktokshop_areas"))

    process_disabled = mode == "Stok Area" and (not areas or not chosen_areas)

    if st.button("Proses", key="btn_stock_tiktokshop", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, issues_bytes, summary = process_stock_tiktokshop(mass_files, pricelist_file, mode, chosen_areas)
            cache_downloads("stock_tiktokshop", "hasil_update_stok_tiktokshop.xlsx", result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_tiktokshop")
    render_downloads("stock_tiktokshop")


def render_harga_normal_shopee():
    page_header(
        "Harga Normal Shopee (Mall & Star)",
        "Mengubah harga normal Shopee (Mall & Star) berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update Shopee (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_shopee_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_shopee_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_shopee_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_shopee_disc")

    if st.button("Proses", key="btn_normal_shopee"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = process_shopee_price_files(
                mass_files, pricelist_file, addon_file, discount_rp, "M4", "Harga Normal Shopee", "normal"
            )
            cache_downloads("normal_shopee", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_shopee")
    render_downloads("normal_shopee")


def render_harga_coret_shopee():
    page_header(
        "Harga Coret Shopee (Mall & Star)",
        "Mengubah harga coret Shopee (Mall & Star) berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Discount Nominate Shopee (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Template Mass Update", type=["xlsx"], accept_multiple_files=True, key="coret_shopee_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="coret_shopee_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="coret_shopee_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="coret_shopee_disc")

    if st.button("Proses", key="btn_coret_shopee"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = process_shopee_price_files(
                mass_files, pricelist_file, addon_file, discount_rp, "M4", "Harga Coret Shopee", "coret"
            )
            cache_downloads("coret_shopee", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("coret_shopee")
    render_downloads("coret_shopee")


def render_harga_normal_tiktokshop():
    page_header(
        "Harga Normal TikTokShop",
        "Mengubah harga normal TikTokShop berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update TikTokShop (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_tiktokshop_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_tiktokshop_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_tiktokshop_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_tiktokshop_disc")

    if st.button("Proses", key="btn_normal_tiktokshop"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = process_tiktokshop_price_normal(
                mass_files, pricelist_file, addon_file, discount_rp
            )
            cache_downloads("normal_tiktokshop", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_tiktokshop")
    render_downloads("normal_tiktokshop")


def render_harga_coret_tiktokshop():
    page_header(
        "Harga Coret TikTokShop",
        "Membuat template output promo TikTokShop berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Input File TikTokShop (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        input_file = st.file_uploader("Upload File TikTokShop", type=["xlsx"], key="coret_tiktokshop_input")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="coret_tiktokshop_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="coret_tiktokshop_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="coret_tiktokshop_disc")

    if st.button("Proses", key="btn_coret_tiktokshop"):
        if not input_file or not pricelist_file or not addon_file:
            st.error("Upload semua file yang dibutuhkan dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = process_tiktokshop_price_coret(
                input_file, pricelist_file, addon_file, discount_rp, True
            )
            cache_downloads("coret_tiktokshop", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("coret_tiktokshop")
    render_downloads("coret_tiktokshop")


def render_harga_normal_powemerchant():
    page_header(
        "Harga Normal PowerMerchant",
        "Mengubah harga normal PowerMerchant berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update PowerMerchant (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_pm_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_pm_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_pm_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_pm_disc")

    if st.button("Proses", key="btn_normal_pm"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = process_powemerchant_price_files(
                mass_files, pricelist_file, addon_file, discount_rp, "Harga Normal PowerMerchant"
            )
            cache_downloads("normal_pm", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_pm")
    render_downloads("normal_pm")


def render_harga_coret_powemerchant():
    page_header(
        "Harga Coret PowerMerchant",
        "Mengubah harga coret PowerMerchant berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update PowerMerchant (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="coret_pm_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="coret_pm_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="coret_pm_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="coret_pm_disc")

    if st.button("Proses", key="btn_coret_pm"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = process_powemerchant_price_files(
                mass_files, pricelist_file, addon_file, discount_rp, "Harga Coret PowerMerchant"
            )
            cache_downloads("coret_pm", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("coret_pm")
    render_downloads("coret_pm")


def render_harga_normal_bigseller():
    page_header(
        "Harga Normal Bigseller",
        "Mengubah harga Bigseller, hanya output baris yang berubah, dan otomatis split 10.000 row per file.",
        ["Mass Update Bigseller (.xlsx, bisa banyak)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_bigseller_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_bigseller_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_bigseller_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_bigseller_disc")

    if st.button("Proses", key="btn_normal_bigseller"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = process_bigseller(
                mass_files, pricelist_file, addon_file, discount_rp
            )
            cache_downloads("normal_bigseller", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_bigseller")
    render_downloads("normal_bigseller")


def render_submit_campaign_shopee():
    page_header(
        "Submit Campaign Shopee",
        "Fitur submit campaign untuk Shopee sedang disiapkan.",
        ["Coming Soon"],
    )
    st.info("Coming Soon")


def render_submit_campaign_tiktokshop():
    page_header(
        "Submit Campaign TikTokShop",
        "Memproses file submit campaign TikTokShop. Output hanya mengambil row yang berubah.",
        [
            "Template Campaign Tiktokshop (.xlsx)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
            "Addon Mapping (.xlsx)",
        ],
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader(
            "Upload Template Campaign Tiktokshop",
            type=["xlsx"],
            accept_multiple_files=True,
            key="submit_campaign_tiktokshop_mass",
        )
    with c2:
        pricelist_file = st.file_uploader(
            "Upload Pricelist",
            type=["xlsx"],
            key="submit_campaign_tiktokshop_pl",
        )
    with c3:
        addon_file = st.file_uploader(
            "Upload Addon Mapping",
            type=["xlsx"],
            key="submit_campaign_tiktokshop_add",
        )

    discount_rp = st.number_input(
        "Diskon (Rp)",
        min_value=0,
        value=0,
        step=1000,
        key="submit_campaign_tiktokshop_disc",
    )

    price_key = st.radio(
        "Ambil harga dari Pricelist",
        ["M3", "M4"],
        horizontal=True,
        key="submit_campaign_tiktokshop_price_key",
    )

    if st.button("Proses", key="btn_submit_campaign_tiktokshop"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return

        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return

        try:
            result_bytes, result_name, issues_bytes, summary = process_submit_campaign_tiktokshop(
                mass_files=mass_files,
                pricelist_file=pricelist_file,
                addon_file=addon_file,
                discount_rp=discount_rp,
                price_key=price_key,
            )
            cache_downloads(
                "submit_campaign_tiktokshop",
                result_name,
                result_bytes,
                issues_bytes,
                summary=summary,
            )
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("submit_campaign_tiktokshop")
    render_downloads("submit_campaign_tiktokshop")


# ============================================================
# SIDEBAR ROUTER
# ============================================================
def build_menu() -> str:
    st.sidebar.title(APP_TITLE)

    group = st.sidebar.radio(
        "Menu Utama",
        ["Dashboard", "Update Stok", "Update Harga Normal", "Update Harga Coret", "Submit Campaign"],
        key="sidebar_main_menu",
    )

    if group == "Dashboard":
        route = "dashboard"

    elif group == "Update Stok":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee (Mall & Star)", "TikTokShop"],
            key="sidebar_update_stok_menu",
        )
        route = "update_stok_shopee" if child.startswith("Shopee") else "update_stok_tiktokshop"

    elif group == "Update Harga Normal":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee (Mall & Star)", "TikTokShop", "PowerMerchant", "Bigseller"],
            key="sidebar_harga_normal_menu",
        )
        if child.startswith("Shopee"):
            route = "harga_normal_shopee"
        elif child == "TikTokShop":
            route = "harga_normal_tiktokshop"
        elif child == "PowerMerchant":
            route = "harga_normal_powermerchant"
        else:
            route = "harga_normal_bigseller"

    elif group == "Update Harga Coret":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee (Mall & Star)", "TikTokShop", "PowerMerchant"],
            key="sidebar_harga_coret_menu",
        )
        if child.startswith("Shopee"):
            route = "harga_coret_shopee"
        elif child == "TikTokShop":
            route = "harga_coret_tiktokshop"
        else:
            route = "harga_coret_powermerchant"

    elif group == "Submit Campaign":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee", "TikTokShop"],
            key="sidebar_submit_campaign_menu",
        )
        if child == "Shopee":
            route = "submit_campaign_shopee"
        else:
            route = "submit_campaign_tiktokshop"

    else:
        route = "dashboard"

    st.sidebar.markdown("---")
    st.sidebar.markdown("<br><br><br>", unsafe_allow_html=True)
    st.sidebar.link_button(
        "Download File Addon",
        "https://drive.google.com/drive/u/0/folders/1r3qVqmm1ALfLGaLuvagAf5EQuMVT0iWI",
        use_container_width=True,
    )

    return route


def main():
    route = build_menu()
    if route == "dashboard":
        render_dashboard()
    elif route == "update_stok_shopee":
        render_update_stok_shopee()
    elif route == "update_stok_tiktokshop":
        render_update_stok_tiktokshop()
    elif route == "harga_normal_shopee":
        render_harga_normal_shopee()
    elif route == "harga_normal_tiktokshop":
        render_harga_normal_tiktokshop()
    elif route == "harga_normal_powermerchant":
        render_harga_normal_powemerchant()
    elif route == "harga_normal_bigseller":
        render_harga_normal_bigseller()
    elif route == "harga_coret_shopee":
        render_harga_coret_shopee()
    elif route == "harga_coret_tiktokshop":
        render_harga_coret_tiktokshop()
    elif route == "harga_coret_powermerchant":
        render_harga_coret_powemerchant()
    elif route == "submit_campaign_shopee":
        render_submit_campaign_shopee()
    elif route == "submit_campaign_tiktokshop":
        render_submit_campaign_tiktokshop()
    else:
        st.error("Menu tidak dikenal.")


if __name__ == "__main__":
    main()
